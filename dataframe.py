from selenium.webdriver.chrome.service import Service
from selenium import webdriver
  
from bs4 import BeautifulSoup as bs
import time
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

# # index=["","12"]
# # new_row_data = [
# #     ['s', 'm', 'L', 'EL'],
# #     ['odigosou', 'dromou', 'dromologio', 'ora']
# #     ]
# # for i in range(len(new_row_data)):
# #     new_row_data[i].insert(0,index[i])
# # print(new_row_data)
# wb = load_workbook("test.xlsx")
# # Select First Worksheet
# ws = wb.worksheets[0]


# # Append 2 new Rows - Columns A - D
# # for row_data in new_row_data:
# #     # Append Row Values
# #     ws.append(row_data)
    
# data=( 
# ("Product","Cost Price","Selling Price"), 
# ("earpod",90, 50), 
# ("laptop", 3000, 8200), 
# ("smartphone", 5100, 7200) 
# )
# for i in data: 
#    ws.append(i)


# wb.save("test.xlsx")

# new dataframe with same columns
#final
df = pd.DataFrame({'Data': [10, 20, 30]})
df.to_excel('test.xlsx', sheet_name='Sheet1', index=False) #saving initial dataframe to file


df1 = pd.DataFrame({'Data': [100, 200, 300]}) # new data

wb = openpyxl.load_workbook('test.xlsx') # open old file
ws = wb["Sheet1"] # assign sheet to work with or as below
# ws = wb.active

for index, row in df1.iterrows():
    ws.append(row.values.tolist())

wb.save("test.xlsx")
